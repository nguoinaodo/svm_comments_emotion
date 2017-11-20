from openpyxl import load_workbook
import csv
import re

def convert_vn_emolex(dirname, filename):
	# Load workbook
	wb = load_workbook('%s%s' % (dirname, filename))
	# Sheet
	ws = wb['Sheet1']
	# Write to csv
	file = open('%s%s' % (dirname, 'vnemolex.csv'), 'w')
	writer = csv.writer(file, delimiter=',')
	# Headers
	cols = ['english', 'vietnamese', 'positive', 'negative', \
			'anger', 'anticipation', 'disgust', 'fear', \
			'joy', 'sadness', 'suprise', 'trust', 'total']
	writer.writerow(cols)
	# Rows
	reg = re.compile('\s')
	for row in ws.iter_rows(min_row=2, min_col=1, max_col=12):
		row_arr = []
		row_arr.append(row[0].value) # english
		vi = re.sub(reg, '_', row[1].value)
		row_arr.append(vi) # vietnamese	
		for cell in row[2:]:
			try:
				row_arr.append(int(cell.value))
			except TypeError:
				row_arr.append(0)
		row_arr.append(sum(row_arr[2:]))
		row_arr = [unicode(x).encode('utf-8') for x in row_arr]
		writer.writerow(row_arr)
	# Close
	file.close()	
		
# Read vietnamese emotion lexicon
def read_vnemolex(filename):
	# Open file
	file = open(filename, 'r')
	# Read
	reader = csv.reader(file, delimiter=',')
	next(reader, None)
	dic = {}
	for row in reader:
		dic[row[1]] = (int(row[2]) + int(row[3])) * 5
	# Close file
	file.close()
	return dic
